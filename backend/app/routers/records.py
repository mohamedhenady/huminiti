from fastapi import APIRouter, HTTPException, UploadFile, File
from app.database import get_supabase
from app.models.schemas import RecordCreate, RecordStatusUpdate
import openpyxl
import io

router = APIRouter(prefix="/records", tags=["records"])

VALID_STATUSES = {"ordered", "prepared", "ready"}


@router.post("/", status_code=201)
async def create_record(data: RecordCreate):
    sb = get_supabase()
    if data.status not in VALID_STATUSES:
        raise HTTPException(status_code=400, detail="Invalid status")
    res = sb.table("records").insert(data.model_dump()).execute()
    return res.data[0]


@router.patch("/{record_id}/status")
async def update_record_status(record_id: str, data: RecordStatusUpdate):
    if data.status not in VALID_STATUSES:
        raise HTTPException(status_code=400, detail="Invalid status. Must be: ordered | prepared | ready")
    sb = get_supabase()
    res = sb.table("records").update({"status": data.status}).eq("id", record_id).execute()
    if not res.data:
        raise HTTPException(status_code=404, detail="Record not found")
    return res.data[0]


@router.delete("/{record_id}", status_code=204)
async def delete_record(record_id: str):
    sb = get_supabase()
    sb.table("records").delete().eq("id", record_id).execute()


@router.post("/import-excel/{batch_id}", status_code=201)
async def import_excel(batch_id: str, file: UploadFile = File(...)):
    """
    Import persons + their drug records from an Excel file.
    Expected columns: full_name, ph_number, location, inv_number, drug_ar_name, drug_en_name, drug_price, is_local
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx/.xls files accepted")

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    ws = wb.active

    headers = [str(cell.value).strip().lower() for cell in ws[1]]
    sb = get_supabase()
    created_persons = {}
    records_to_insert = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))
        if not row_data.get("full_name"):
            continue

        name = str(row_data["full_name"]).strip()
        if name not in created_persons:
            person_payload = {
                "full_name": name,
                "ph_number": str(row_data.get("ph_number", "") or ""),
                "location": str(row_data.get("location", "") or ""),
                "inv_number": str(row_data.get("inv_number", "") or ""),
                "batch_id": batch_id,
            }
            p_res = sb.table("persons").insert(person_payload).execute()
            created_persons[name] = p_res.data[0]["id"]

        person_id = created_persons[name]

        # Look up or create drug
        drug_ar = str(row_data.get("drug_ar_name", "") or "").strip()
        if not drug_ar:
            continue
        price = float(row_data.get("drug_price", 0) or 0)
        is_local = str(row_data.get("is_local", "true")).lower() in ("true", "yes", "1", "محلي")
        discount = 0.80 if is_local else 0.90
        final_price = round(price * discount, 2)

        drug_res = sb.table("drugs").select("id").eq("ar_name", drug_ar).execute()
        if drug_res.data:
            drug_id = drug_res.data[0]["id"]
        else:
            new_drug = sb.table("drugs").insert({
                "ar_name": drug_ar,
                "en_name": str(row_data.get("drug_en_name", "") or ""),
                "price": price,
                "is_local": is_local,
            }).execute()
            drug_id = new_drug.data[0]["id"]

        records_to_insert.append({
            "person_id": person_id,
            "batch_id": batch_id,
            "drug_id": drug_id,
            "drug_name": drug_ar,
            "drug_price": price,
            "final_price": final_price,
            "status": "ordered",
        })

    if records_to_insert:
        sb.table("records").insert(records_to_insert).execute()

    return {
        "persons_created": len(created_persons),
        "records_created": len(records_to_insert),
    }
